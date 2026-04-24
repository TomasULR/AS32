"""Generate the SA32 v1.1 Bill of Materials as an XLSX file.

v1.1 změny vs. Phase 1:
  + BK3266 BT receiver (externí I²S, řízený přes UART AT)
  + NX3L4684 / 74HC4053 I²S MUX
  + IRM-45-24 PSU (1,87 A místo 0,83 A)
  + HLK-5M05 izolovaný buck 24V→5V
  + Externí 3,3 V LDO (uvolní DevKit AMS1117)
  + Ferritový bead BLM31KN601 + filtr kondenzátory

Ceny: typické CZ retail 2026 (CZK, jednokusové). Slouží pro rozpočet,
ne jako oficiální nabídka.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/tomasurl/AS32/AS32/docs/SA32_BOM.xlsx"

# (kategorie, qty, ref, part, package/notes, e-shop, url, est_unit_price_czk)
COMPONENTS = [
    # ---- Core compute ----
    ("MCU", 1, "U1",
     "ESP32-S3-DevKitC-1 N16R8 (16 MB flash, 8 MB octal PSRAM)",
     "DevKit modul s USB-C; ⚠ ověř N16R8 variantu, ne N8R8",
     "Allegro.cz / LaskaKit",
     "https://allegro.cz/oferta/esp32-s3-devkitc-1-n16r8-wroom-1",
     316),

    # ---- Audio chain ----
    ("Audio", 1, "U2",
     "PCM5102A I²S DAC modul (GY-PCM5102, line-out 2,1 Vrms)",
     "modul ~38×25 mm, SCK→GND interní PLL", "LaskaKit",
     "https://www.laskakit.cz/gy-pcm5102-i2s-iis-stereo-dac-modul/",
     130),
    ("Audio", 1, "U3",
     "TPA3110D2 2×15 W class-D modul (XH-A232)",
     "PBTL/stereo jumper, VIN 12–24 V; ⚠ nastav GAIN=12 dB jumperem",
     "Hadex",
     "https://www.hadex.cz/m487-zesilovac-2x15w-tpa3110-modul-stereo/",
     120),
    ("Audio", 1, "U7",
     "BK3266 / BLK-MD-SPK Bluetooth receiver s I²S výstupem",
     "BT 5.0 BR/EDR, A2DP/AVRCP přes UART AT příkazy, 3,3 V napájení",
     "Allegro.cz",
     "https://allegro.cz/oferta/bluetooth-5-0-modul-bk3266-i2s-receiver",
     130),
    ("Audio", 1, "U8",
     "I²S MUX NX3L4684 (3-kanál 2:1, R_on 0,5 Ω)",
     "TSSOP-14; alt. 74HC4053 v DIP-16 (8 Kč) pro DIY",
     "TME.cz",
     "https://www.tme.eu/cz/details/nx3l4684tk/multiplexery-cmos/nexperia/",
     30),
    ("Audio", 1, "J1",
     "Stereo 3,5 mm jack PCB pro propojení DAC → AMP",
     "stíněný kablík ~10 cm, stínění zem jen na DAC straně",
     "GME",
     "https://www.gme.cz/v/1503180/jack-konektor-3-5-stereo-do-dps",
     35),

    # ---- UI ----
    ("UI", 1, "U4",
     "OLED displej SSD1306 / SSD1315 0,96\" 128×64 I²C",
     "I²C 0x3C, napájení 3,3 V (NIKDY ne 5 V), pull-upy onboard",
     "LaskaKit",
     "https://www.laskakit.cz/oled-displej-i2c-128x64-0-96--bily/",
     100),
    ("UI", 1, "S1",
     "Rotační enkodér Alps EC11 (24 pulzů, taktilní switch)",
     "vyveden A/B/SW + 2× GND; alt. 20-pulzní za 41 Kč na Allegru",
     "TME.cz",
     "https://www.tme.eu/cz/details/ec11k-15-15a/enkodery/bourns/",
     70),
    ("UI", 1, "K1",
     "Hliníkový knoflík ⌀20 mm, D-shaft 6 mm",
     "matný hliník, hřídel 20 mm",
     "TME.cz",
     "https://www.tme.eu/cz/details/g-pkn1810b/knofliky/",
     60),

    # ---- Power ----
    ("Power", 1, "PSU",
     "Mean Well IRM-45-24 (24 V / 1,88 A / 45 W, AC/DC zalitý modul)",
     "drop-in upgrade z IRM-20-24, +30 % rezerva nad 1,44 A peak AMP",
     "GME",
     "https://www.gme.cz/v/1499847/irm-45-24-mean-well-spinany-zdroj",
     490),
    ("Power", 1, "U6",
     "HLK-5M05 izolovaný buck 24 V → 5 V / 1 A (5 W)",
     "primárně izolovaný = bez ground loopu; alt. MP1584 (45 Kč) ale rušivý",
     "GME",
     "https://www.gme.cz/v/1505423/hlk-5m05-spinany-zdroj-5v-1a",
     90),
    ("Power", 1, "U9",
     "AMS1117-3.3 / AP2112K-3.3 modul nebo SOT-223 + caps",
     "5 V → 3,3 V LDO pro DAC + OLED (~30 mA), uvolní DevKit AMS1117",
     "GME",
     "https://www.gme.cz/v/1499432/ams1117-3-3-stabilizator-3-3v-sot-223",
     25),
    ("Power", 1, "F1",
     "Pojistka 5×20 mm 1,6 A T (slow-blow) + držák do panelu",
     "primární strana 230 V — POVINNÉ pro bezpečnost",
     "GME",
     "https://www.gme.cz/v/1499143/pojistka-5x20-1-6a-t",
     35),
    ("Power", 1, "J2",
     "DC jack 5,5 / 2,1 mm panelový (pro IRM-45-24 výstup)",
     "nebo přímý drát z IRM do AMP (úspora 30 Kč, mínus modularita)",
     "GME",
     "https://www.gme.cz/v/1499064/dc-jack-5-5-2-1-do-panelu",
     30),

    # ---- Pasivní ----
    ("Passive", 2, "C1, C2",
     "Elektrolyt 1000 µF / 35 V low-ESR",
     "filtrace 24 V vstupu AMP; doporučeno Panasonic FM serie", "GME",
     "https://www.gme.cz/v/1500253/c-1000u-35v-rad-105-fm",
     25),
    ("Passive", 4, "C3-C6",
     "MLCC 100 nF / 50 V X7R",
     "blok napájení každého aktivního IC (DAC, MUX, BT, LDO)", "GME",
     "https://www.gme.cz/v/1505200/cm-100n-50-x7r",
     3),
    ("Passive", 2, "C7, C8",
     "MLCC 47 µF / 16 V X5R",
     "Pi filtr na 5 V výstupu HLK-5M05 (před a za ferrit beadem)", "GME",
     "https://www.gme.cz/v/1502401/cm-47u-16v-x5r",
     8),
    ("Passive", 1, "L1",
     "Ferritový bead BLM31KN601SN1 (600 Ω @ 100 MHz, 0805/1206)",
     "v sérii s 5 V z buck před AMS1117 — potlačí spínací harmonické", "GME",
     "https://www.gme.cz/v/1505421/blm31-600r",
     5),
    ("Passive", 4, "R1-R4",
     "Sada rezistorů (4,7 kΩ pull-up, 33 Ω I²C série, 470 Ω LED)",
     "1/4 W, 5 %; pokud máš sadu 600+ ks, použij ji", "GME",
     "https://www.gme.cz/v/sada-rezistoru",
     2),

    # ---- Mechanika ----
    ("Mech", 2, "SP",
     "Reproduktor 8 Ω / ≥ 15 W, 3–4\" full-range",
     "podle krabičky a poslechu; alternativa 2-pásmový mini-monitor",
     "TME.cz",
     "https://www.tme.eu/cz/details/dl-110-8/reproduktory-9/",
     280),
    ("Mech", 1, "BOX",
     "Hliníková krabička Hammond 1455N1601 (~120×100×45 mm)",
     "frézování čela pro OLED + enkodér; alt. levnější ABS box ~200 Kč",
     "TME.cz",
     "https://www.tme.eu/cz/details/1455n1601/standardni-pouzdra-instrumentalni/hammond/",
     450),
    ("Mech", 1, "WIRE",
     "Sada propojovacích vodičů Dupont F-F + silnější 18 AWG",
     "mix barev pro snadnou orientaci",
     "GME",
     "https://www.gme.cz/v/sada-vodicu-dupont",
     80),
    ("Mech", 1, "STAND",
     "Distanční sloupky M3 mosaz 8 mm + šrouby/matky",
     "izolace modulů od šasi, sada ~20 ks",
     "GME",
     "https://www.gme.cz/v/distancni-sloupek-m3-mosaz",
     45),

    # ---- Volitelné ----
    ("Optional", 2, "LED1, LED2",
     "Status LED 3 mm + 470 Ω (Wi-Fi / BT indikace na GPIO 38/39)",
     "Kconfig SOKOL_STATUS_LEDS_ENABLE",
     "GME",
     "https://www.gme.cz/v/led-3mm",
     8),
    ("Optional", 1, "U10",
     "Krátký USB-C kabel pro flashování + sériovou konzoli",
     "USB-C k DevKitu, 0,5 m stačí",
     "GME",
     "https://www.gme.cz/v/usb-c-kabel",
     90),
]

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
HEADER_FILL = PatternFill("solid", fgColor="0D9488")
SECTION_FILL = PatternFill("solid", fgColor="E2E8F0")
TOTAL_FILL = PatternFill("solid", fgColor="1D4ED8")
NEW_FILL = PatternFill("solid", fgColor="FEF3C7")  # zvýraznění v1.1 přidaného

V1_1_NEW = {"U7", "U8", "U6", "U9", "L1", "C7, C8"}  # nové v v1.1
V1_1_UPGRADED = {"PSU"}                                # upgrade existujícího

def main():
    wb = Workbook()

    # ---- Sheet 1: BOM ----
    ws = wb.active
    ws.title = "BOM v1.1"
    headers = ["#", "Kategorie", "Ks", "Ozn.", "Součástka",
               "Pouzdro / poznámka", "E-shop", "URL",
               "Cena/ks (CZK)", "Cena celkem (CZK)", "v1.1"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    total = 0
    for i, (cat, qty, ref, part, pkg, shop, url, price) in enumerate(COMPONENTS, start=1):
        line_total = qty * price
        if cat != "Optional":
            total += line_total
        marker = ""
        fill = None
        if ref in V1_1_NEW:        marker, fill = "NEW",     NEW_FILL
        elif ref in V1_1_UPGRADED: marker, fill = "UPGRADE", NEW_FILL
        ws.append([i, cat, qty, ref, part, pkg, shop, url, price, line_total, marker])

        for col in range(1, len(headers) + 1):
            c = ws.cell(row=i + 1, column=col)
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if cat == "Optional":
                c.fill = SECTION_FILL
            elif fill is not None:
                c.fill = fill
            if col == 8:  # URL sloupec → hyperlink
                if url:
                    c.hyperlink = url
                    c.font = Font(color="1D4ED8", underline="single", size=9)
                    c.value = "otevřít"

    last = ws.max_row
    ws.append([])
    ws.cell(row=last + 2, column=9, value="CELKEM (povinné, orient.)").font = Font(bold=True)
    total_cell = ws.cell(row=last + 2, column=10, value=total)
    total_cell.font = Font(bold=True, color="FFFFFF")
    total_cell.fill = TOTAL_FILL
    total_cell.border = BORDER

    widths = [4, 11, 4, 9, 42, 38, 18, 12, 12, 14, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # ---- Sheet 2: GPIO map (v1.1) ----
    g = wb.create_sheet("GPIO map v1.1")
    gh = ["GPIO", "Funkce", "Směr", "Pull / IRQ", "Připojení", "Pozn."]
    g.append(gh)
    for col in range(1, len(gh) + 1):
        c = g.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    GPIO = [
        (5,  "I²S BCLK",     "out", "—",                "MUX → PCM5102A BCK", "1,4112 MHz @ 44,1 kHz S16"),
        (6,  "I²S LRCK",     "out", "—",                "MUX → PCM5102A LCK", "44,1 kHz word select"),
        (7,  "I²S DOUT",     "out", "—",                "MUX → PCM5102A DIN", "stereo S16LE"),
        (15, "I²S MCLK (opt)","out","—",                "PCM5102A SCK",       "Kconfig SOKOL_DAC_MCLK_ENABLE"),
        (8,  "I²C SDA",      "io",  "PU int.",          "SSD1306 SDA",        "400 kHz"),
        (9,  "I²C SCL",      "out", "PU int.",          "SSD1306 SCL",        "400 kHz"),
        (10, "Encoder A",    "in",  "PU + IRQ any-edge","EC11 A",             "FSM v rotary_ec11.c"),
        (11, "Encoder B",    "in",  "PU + IRQ any-edge","EC11 B",             ""),
        (12, "Encoder SW",   "in",  "PU + IRQ falling", "EC11 SW",            "active LOW, debounce SW"),
        (21, "AMP MUTE",     "out", "—",                "TPA3110D2 MUTE",     "active HIGH = ticho, default boot"),
        (13, "DAC XSMT (opt)","out","—",                "PCM5102A XSMT",      "Kconfig SOKOL_DAC_MUTE_PIN_ENABLE"),
        (4,  "I²S MUX SEL",  "out", "PD external",      "NX3L4684 SEL",       "v1.1 NEW: 0=ESP, 1=BT"),
        (14, "BK3266 RESET", "out", "—",                "BK3266 nReset",      "v1.1 NEW: active LOW pulse"),
        (17, "UART2 TX",     "out", "—",                "BK3266 RX",          "v1.1 NEW: AT příkazy 115200"),
        (18, "UART2 RX",     "in",  "—",                "BK3266 TX",          "v1.1 NEW: status events"),
        (38, "Wi-Fi LED (opt)","out","—",               "LED → 470 Ω → GND",  "Kconfig SOKOL_STATUS_LEDS_ENABLE"),
        (39, "BT LED (opt)", "out", "—",                "LED → 470 Ω → GND",  ""),
    ]
    for row in GPIO:
        g.append(row)
        r = g.max_row
        for col in range(1, len(gh) + 1):
            g.cell(row=r, column=col).border = BORDER
            g.cell(row=r, column=col).alignment = Alignment(vertical="top", wrap_text=True)
            if "v1.1 NEW" in str(row[5]):
                g.cell(row=r, column=col).fill = NEW_FILL
    for i, w in enumerate([7, 16, 6, 18, 22, 42], start=1):
        g.column_dimensions[get_column_letter(i)].width = w
    g.row_dimensions[1].height = 24
    g.freeze_panes = "A2"

    # ---- Sheet 3: Power budget ----
    p = wb.create_sheet("Napájení v1.1")
    ph = ["Větev", "Napětí", "Typický odběr", "Špička", "Zdroj", "Ztráta"]
    p.append(ph)
    for col in range(1, len(ph) + 1):
        c = p.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    PWR = [
        ("AMP TPA3110D2 (2×15 W RMS)", "24 V", "150 mA klid",  "1,44 A peak",       "IRM-45-24 (1,88 A) — 30 % rezerva", "4,5 W na čipu (heatsink)"),
        ("Logic 5 V (přes HLK-5M05)",  "24 V", "0,2 A vstup",  "0,5 A vstup",       "IRM-45-24 stejný rail",            "0,33 W v HLK"),
        ("DevKit ESP32-S3 (přes onb. AMS1117)", "5 V", "120 mA",   "500 mA Wi-Fi TX", "HLK-5M05",                         "0,93 W na DevKit AMS"),
        ("BK3266 BT receiver",         "3,3 V","35 mA",        "50 mA streaming",   "externí AMS1117 (U9) z 5 V",       "<0,1 W"),
        ("PCM5102A DAC",               "3,3 V","20 mA",        "30 mA",             "externí AMS1117 (U9) z 5 V",       "<0,1 W"),
        ("SSD1306 OLED",               "3,3 V","8 mA",         "20 mA",             "externí AMS1117 (U9) z 5 V",       "<0,1 W"),
        ("EC11 enkodér",               "3,3 V","< 1 mA",       "—",                 "interní pull-upy ESP32",           "<0,01 W"),
        ("NX3L4684 MUX",               "3,3 V","< 1 mA",       "10 mA",             "externí AMS1117 (U9) z 5 V",       "<0,01 W"),
        ("Reproduktory 2× 8 Ω",        "—",    "—",            "2× 15 W RMS",       "BTL z TPA3110D2",                  "—"),
    ]
    for row in PWR:
        p.append(row)
        r = p.max_row
        for col in range(1, len(ph) + 1):
            p.cell(row=r, column=col).border = BORDER
            p.cell(row=r, column=col).alignment = Alignment(vertical="top", wrap_text=True)
    for i, w in enumerate([28, 8, 16, 18, 30, 22], start=1):
        p.column_dimensions[get_column_letter(i)].width = w
    p.row_dimensions[1].height = 24
    p.freeze_panes = "A2"

    # ---- Sheet 4: Validation report ----
    v = wb.create_sheet("Validace HW")
    vh = ["#", "Kontrola", "Spec / norma", "Výsledek", "Pozn."]
    v.append(vh)
    for col in range(1, len(vh) + 1):
        c = v.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    CHECKS = [
        (1, "PSU nadproudová rezerva",    "I_max ≥ 1,3 × I_peak", "OK", "1,88 A / 1,44 A = 130 %"),
        (2, "5 V buck výkonová rezerva",  "P_max ≥ 1,5 × P_avg",   "OK", "5 W / 2,8 W = 178 %"),
        (3, "AMS1117 termální (DevKit)",  "P_diss < 1 W bez heatsinku", "BORDERLINE", "0,93 W peak — odlehčit ext. LDO"),
        (4, "I²S BCLK přes MUX",          "f_max NX3L4684 > 1,4 MHz", "OK", "NX3L4684 propustnost > 100 MHz"),
        (5, "Logické úrovně I²S",         "V_OH ≥ 2,0 V (V_IH PCM5102A)", "OK", "ESP/BT 3,0 V min, DAC 2,0 V"),
        (6, "DAC → AMP voltage match",    "V_in ≤ 1 Vrms při GAIN 20 dB", "ACTION", "Nastavit GAIN jumper na 12 dB"),
        (7, "Pojistka primár",            "I_t = 1,6 A T pro 45 W zdroj", "OK", "F1 v sérii s L 230 V"),
        (8, "BK3266 baud rate",           "115200 default firmware", "OK", "potvrzeno pro většinu klonů"),
        (9, "Partition table OTA",        "2× ota_X (2 MB) + factory", "OK", "partitions.csv beze změny"),
        (10,"Disable BR/EDR (ESP32-S3)",  "ESP_BT_ENABLED=n",       "OK", "úspora ~80 KB internal RAM"),
        (11,"GPIO konflikty",             "žádné s octal PSRAM 35-37", "OK", "MUX_SEL 4, BK_RST 14, UART2 17/18"),
        (12,"Audio cable shield",         "ground only at DAC end", "DOC", "v PDF zmíněno, vizuální kontrola"),
        (13,"Star ground topologie",      "single junction at PSU GND", "DOC", "manuál instaluje uživatel"),
        (14,"OTA rollback",               "BOOTLOADER_APP_ROLLBACK_ENABLE", "OK", "auto-validate po 30 s běhu"),
        (15,"HTTPS cert verify",          "MBEDTLS_CERTIFICATE_BUNDLE",   "OK", "mozilla CA bundle"),
    ]
    for row in CHECKS:
        v.append(row)
        r = v.max_row
        for col in range(1, len(vh) + 1):
            cell = v.cell(row=r, column=col)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col == 4:
                txt = str(row[3])
                if txt == "OK":
                    cell.font = Font(bold=True, color="047857")
                elif txt in ("BORDERLINE", "ACTION"):
                    cell.font = Font(bold=True, color="B45309")
                elif txt == "DOC":
                    cell.font = Font(bold=True, color="1D4ED8")
    for i, w in enumerate([4, 28, 30, 14, 38], start=1):
        v.column_dimensions[get_column_letter(i)].width = w
    v.row_dimensions[1].height = 24
    v.freeze_panes = "A2"

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"  Total (povinné): {total} CZK")

if __name__ == "__main__":
    main()
